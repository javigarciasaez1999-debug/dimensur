from __future__ import annotations

import logging
import unicodedata
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Iterable

import gspread
from gspread.exceptions import APIError, SpreadsheetNotFound, WorksheetNotFound
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .models import SheetRow


HEADERS = [
    "ID",
    "Título base",
    "Keyword principal",
    "Estado",
    "Contenido HTML",
    "Título SEO",
    "Metadescripción",
    "Slug",
    "Categoría",
    "Etiquetas",
    "Enlaces internos usados",
    "Anchors usados",
    "Resumen del contenido generado",
    "Diferencia frente al artículo anterior",
    "Prompt imagen",
    "Concepto visual imagen",
    "Diferencia visual frente a imagen anterior",
    "URL imagen o nombre archivo",
    "ALT imagen",
    "JSON enviado",
    "Respuesta API",
    "URL publicada",
    "Fecha publicación",
    "Error",
]


def _header_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_value = "".join(
        char for char in normalized if not unicodedata.combining(char)
    )
    return " ".join(ascii_value.lower().strip().split())


CANONICAL_BY_KEY = {_header_key(item): item for item in HEADERS}
CANONICAL_BY_KEY.update(
    {
        "titulos de contenidos": "Título base",
        "contenido": "Contenido HTML",
        "titulo": "Título SEO",
        "metadescripcion": "Metadescripción",
        "imagen a usar": "URL imagen o nombre archivo",
    }
)


class SheetConfigurationError(RuntimeError):
    pass


class BaseSheetsClient(ABC):
    def __init__(self, logger: logging.Logger) -> None:
        self.logger = logger
        self._headers: list[str] = []
        self._column_by_header: dict[str, int] = {}

    @abstractmethod
    def load_sheet(self) -> None:
        raise NotImplementedError

    @abstractmethod
    def _all_values(self) -> list[list[Any]]:
        raise NotImplementedError

    @abstractmethod
    def _write_fields(self, row_number: int, fields: dict[str, Any]) -> None:
        raise NotImplementedError

    def get_headers(self) -> list[str]:
        return list(self._headers)

    def missing_headers(self) -> list[str]:
        return [header for header in HEADERS if header not in self._column_by_header]

    def require_headers(self) -> None:
        missing = self.missing_headers()
        if missing:
            raise SheetConfigurationError(
                "Faltan columnas obligatorias en la hoja: " + ", ".join(missing)
            )

    def _append_missing_headers(self, raw_headers: list[Any]) -> list[str]:
        self._refresh_header_map(raw_headers)
        missing = self.missing_headers()
        if not missing:
            return self._headers
        return self._headers + missing

    def _refresh_header_map(self, raw_headers: Iterable[Any]) -> None:
        self._headers = [str(value).strip() for value in raw_headers]
        self._column_by_header = {}
        for index, raw_header in enumerate(self._headers, start=1):
            canonical = CANONICAL_BY_KEY.get(_header_key(raw_header), raw_header)
            if canonical:
                self._column_by_header[canonical] = index

    def _row_from_values(self, row_number: int, values: list[Any]) -> SheetRow:
        mapped: dict[str, str] = {}
        for header, column in self._column_by_header.items():
            index = column - 1
            mapped[header] = (
                str(values[index]).strip()
                if index < len(values) and values[index] is not None
                else ""
            )
        return SheetRow(row_number=row_number, values=mapped)

    def get_row(self, row_number: int) -> SheetRow:
        values = self._all_values()
        if row_number < 2:
            raise ValueError("El número de fila debe ser 2 o superior.")
        if row_number > len(values):
            raise ValueError(f"La fila {row_number} no existe en la hoja.")
        return self._row_from_values(row_number, values[row_number - 1])

    def find_next_pending_row(
        self, statuses: tuple[str, ...] = ("Crear",)
    ) -> SheetRow | None:
        values = self._all_values()
        status_column = self._column_by_header.get("Estado")
        if not status_column:
            raise SheetConfigurationError("No existe la columna Estado.")
        accepted = set(statuses)
        for row_number, row_values in enumerate(values[1:], start=2):
            status_index = status_column - 1
            status = (
                str(row_values[status_index]).strip()
                if status_index < len(row_values)
                else ""
            )
            if status in accepted:
                return self._row_from_values(row_number, row_values)
        return None

    def update_row_status(self, row_number: int, status: str) -> None:
        self._write_fields(row_number, {"Estado": status})

    def update_row_fields(self, row_number: int, fields: dict[str, Any]) -> None:
        unknown = [field for field in fields if field not in self._column_by_header]
        if unknown:
            raise SheetConfigurationError(
                "No se pueden actualizar columnas inexistentes: " + ", ".join(unknown)
            )
        self._write_fields(row_number, fields)

    def get_last_published_article(
        self, *, exclude_row: int | None = None
    ) -> dict[str, Any] | None:
        values = self._all_values()
        for row_number in range(len(values), 1, -1):
            if exclude_row and row_number == exclude_row:
                continue
            row = self._row_from_values(row_number, values[row_number - 1])
            if row.get("Estado") not in {
                "Publicado",
                "Publicada",
                "Generado - No publicado",
            }:
                continue
            return {
                "source_id": row.get("ID") or f"row-{row_number}",
                "title": row.get("Título base") or row.get("Título SEO"),
                "slug": row.get("Slug"),
                "summary": row.get("Resumen del contenido generado"),
                "html": row.get("Contenido HTML"),
                "internal_links_used": row.get("Enlaces internos usados"),
                "image_concept": row.get("Concepto visual imagen"),
                "difference_vs_previous": row.get(
                    "Diferencia frente al artículo anterior"
                ),
                "row_number": row_number,
            }
        return None


class GoogleSheetsClient(BaseSheetsClient):
    def __init__(
        self,
        spreadsheet_id: str,
        worksheet_name: str,
        credentials_file: Path,
        logger: logging.Logger,
    ) -> None:
        super().__init__(logger)
        self.spreadsheet_id = spreadsheet_id
        self.worksheet_name = worksheet_name
        self.credentials_file = credentials_file
        self.worksheet: gspread.Worksheet | None = None

    def load_sheet(self) -> None:
        if not self.spreadsheet_id:
            raise SheetConfigurationError("GOOGLE_SHEET_URL/GOOGLE_SHEET_ID está vacío.")
        if not self.credentials_file.exists():
            raise SheetConfigurationError(
                f"No existe GOOGLE_CREDENTIALS_FILE: {self.credentials_file}"
            )
        try:
            client = gspread.service_account(filename=str(self.credentials_file))
            spreadsheet = client.open_by_key(self.spreadsheet_id)
            self.worksheet = spreadsheet.worksheet(self.worksheet_name)
            raw_headers = self.worksheet.row_values(1)
            if not any(str(value).strip() for value in raw_headers):
                self.worksheet.update(range_name="A1:X1", values=[HEADERS])
                raw_headers = HEADERS
            else:
                completed_headers = self._append_missing_headers(raw_headers)
                if len(completed_headers) > len(raw_headers):
                    start_column = get_column_letter(len(raw_headers) + 1)
                    end_column = get_column_letter(len(completed_headers))
                    self.worksheet.update(
                        range_name=f"{start_column}1:{end_column}1",
                        values=[completed_headers[len(raw_headers) :]],
                    )
                    raw_headers = completed_headers
            self._refresh_header_map(raw_headers)
        except (SpreadsheetNotFound, WorksheetNotFound, APIError) as exc:
            raise SheetConfigurationError(
                f"No se pudo abrir Google Sheet '{self.worksheet_name}': {exc}"
            ) from exc

    def _require_worksheet(self) -> gspread.Worksheet:
        if self.worksheet is None:
            raise RuntimeError("La hoja no está cargada. Llama a load_sheet().")
        return self.worksheet

    def _all_values(self) -> list[list[Any]]:
        return self._require_worksheet().get_all_values()

    def _write_fields(self, row_number: int, fields: dict[str, Any]) -> None:
        worksheet = self._require_worksheet()
        requests = []
        for header, value in fields.items():
            column = self._column_by_header[header]
            cell = f"{get_column_letter(column)}{row_number}"
            requests.append({"range": cell, "values": [[str(value)]]})
        try:
            worksheet.batch_update(requests, value_input_option="RAW")
        except APIError as exc:
            self.logger.exception("Error escribiendo en Google Sheets")
            raise RuntimeError(f"Error escribiendo en Google Sheets: {exc}") from exc


class ExcelSheetsClient(BaseSheetsClient):
    def __init__(
        self, file_path: Path, worksheet_name: str, logger: logging.Logger
    ) -> None:
        super().__init__(logger)
        self.file_path = file_path
        self.worksheet_name = worksheet_name
        self.workbook = None
        self.worksheet = None

    def load_sheet(self) -> None:
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        if not self.file_path.exists():
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self.worksheet_name
            worksheet.append(HEADERS)
            workbook.save(self.file_path)

        self.workbook = load_workbook(self.file_path)
        if self.worksheet_name not in self.workbook.sheetnames:
            raise SheetConfigurationError(
                f"No existe la hoja Excel '{self.worksheet_name}' en {self.file_path}"
            )
        self.worksheet = self.workbook[self.worksheet_name]
        raw_headers = [
            self.worksheet.cell(row=1, column=column).value or ""
            for column in range(1, self.worksheet.max_column + 1)
        ]
        if not any(str(value).strip() for value in raw_headers):
            for column, header in enumerate(HEADERS, start=1):
                self.worksheet.cell(row=1, column=column, value=header)
            self.workbook.save(self.file_path)
            raw_headers = HEADERS
        else:
            completed_headers = self._append_missing_headers(raw_headers)
            if len(completed_headers) > len(raw_headers):
                for column, header in enumerate(completed_headers, start=1):
                    self.worksheet.cell(row=1, column=column, value=header)
                self.workbook.save(self.file_path)
                raw_headers = completed_headers
        self._refresh_header_map(raw_headers)

    def _require_sheet(self):
        if self.worksheet is None or self.workbook is None:
            raise RuntimeError("El Excel no está cargado. Llama a load_sheet().")
        return self.worksheet

    def _all_values(self) -> list[list[Any]]:
        worksheet = self._require_sheet()
        return [
            [cell.value if cell.value is not None else "" for cell in row]
            for row in worksheet.iter_rows()
        ]

    def _write_fields(self, row_number: int, fields: dict[str, Any]) -> None:
        worksheet = self._require_sheet()
        for header, value in fields.items():
            text = str(value)
            if len(text) > 32_767:
                text = text[:32_640] + "\n...[TRUNCADO POR EL LÍMITE DE CELDA DE EXCEL]"
            worksheet.cell(
                row=row_number, column=self._column_by_header[header], value=text
            )
        try:
            self.workbook.save(self.file_path)
        except OSError as exc:
            self.logger.exception("Error escribiendo el Excel")
            raise RuntimeError(f"Error escribiendo el Excel: {exc}") from exc


def create_sheet_client(settings, logger: logging.Logger) -> BaseSheetsClient:
    if settings.sheet_backend == "excel":
        return ExcelSheetsClient(
            settings.excel_file_path, settings.google_sheet_name, logger
        )
    return GoogleSheetsClient(
        settings.google_sheet_id,
        settings.google_sheet_name,
        settings.google_credentials_file,
        logger,
    )
